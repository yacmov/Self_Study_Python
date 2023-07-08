from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet를 생성 - 기본이름으로 생성
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws1.sheet_properties.tabColor = "fe44ef"

ws2 = wb.create_sheet("NewSheet", 2)

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근이 가능하다

print(wb.sheetnames ) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample.xlsx")
wb.close()

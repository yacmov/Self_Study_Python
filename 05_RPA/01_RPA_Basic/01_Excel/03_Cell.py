from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1 이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

for i in range(1, 25):
    ws[f"C{i}"] = i

print(ws["A1"]) # A1 셀의 객체 정보를 출력
print(ws["A1"].value) # A1 셀의 값 정보를 출력

for i in range(1,25):
    print(ws[f"C{i}"].value)

# row = 1,2,3 ...
# column = A(1), B(2), C(3)....
ws.cell(row=1, column=1)

ws.cell(column=3, row=1, value=10) # 아래와 같은 결과
ws["C1"] = 10 # 위와 같은 결과 

from random import *
for x in range(1,11):
    for y in range(1,11):
        ws.cell(column = x, row = y, value= randint(1,1000))



wb.save("sample.xlsx")
wb.close()
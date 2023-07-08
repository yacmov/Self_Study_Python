from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0,100), randint(0,100) ])

# col_B = ws["B"]
# # print(col_B)
# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"] # 영어, 수학 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1] # 1 번재 row 만 가져오기
# for cell in row_title:
#     print(cell.value) 

# row_range = ws[2:6] # 2 번재 부터 6번째 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ") 
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         #print(cell.value, end=" ") # 각 셀에 벨류 정보를 가져온다
#         print(cell.coordinate, end=" ") # 각 셀에 대한 정보를 가져온다
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()

# 전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value)


# 전체 columns
# print(tuple(ws.rows))
# for column in tuple(ws.colums):
#     print(column[0].value)

# for row in ws.iter_rows(): # 전체 row
#     print(row[1].value)

# for column in ws.iter_cols(): # 전체 colum
#     print(column[0].value)

for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)


wb.save("sample.xlsx")
wb.close()